from __future__ import annotations

import logging
from datetime import datetime

import httpx
from celery import shared_task
from django.conf import settings
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime

from apps.core.models import SystemSettings

from .models import SyncCheckpoint, Vulnerability
from .services import merge_bdu_into_cve, severity_from_score

logger = logging.getLogger(__name__)

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"


def _checkpoint(source: str) -> SyncCheckpoint:
    obj, _ = SyncCheckpoint.objects.get_or_create(source=source)
    return obj


def _parse_dt(value):
    if not value:
        return None
    dt = parse_datetime(value) if isinstance(value, str) else value
    if dt is None and isinstance(value, str):
        d = parse_date(value)
        if d:
            dt = datetime.combine(d, datetime.min.time())
    if dt and timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _extract_cvss(metrics: dict) -> dict:
    out = {"v31": None, "v30": None, "v2": None, "v4": None}
    if not isinstance(metrics, dict):
        return out

    def first_metric(keys):
        for key in keys:
            items = metrics.get(key) or []
            if items:
                return items[0]
        return None

    m31 = first_metric(["cvssMetricV31"])
    if m31:
        data = m31.get("cvssData") or {}
        out["v31"] = {
            "score": data.get("baseScore"),
            "vector": data.get("vectorString"),
            "severity": data.get("baseSeverity"),
            "raw": m31,
        }
    m30 = first_metric(["cvssMetricV30"])
    if m30:
        data = m30.get("cvssData") or {}
        out["v30"] = {
            "score": data.get("baseScore"),
            "vector": data.get("vectorString"),
            "severity": data.get("baseSeverity"),
            "raw": m30,
        }
    m2 = first_metric(["cvssMetricV2"])
    if m2:
        data = m2.get("cvssData") or {}
        out["v2"] = {
            "score": data.get("baseScore"),
            "vector": data.get("vectorString"),
            "severity": m2.get("baseSeverity") or data.get("baseSeverity"),
            "raw": m2,
        }
    m4 = first_metric(["cvssMetricV40", "cvssMetricV4"])
    if m4:
        data = m4.get("cvssData") or {}
        out["v4"] = {
            "score": data.get("baseScore"),
            "vector": data.get("vectorString"),
            "severity": data.get("baseSeverity"),
            "raw": m4,
        }
    return out


def _upsert_nvd_item(item: dict) -> Vulnerability:
    cve = item.get("cve") or item
    vuln_id = cve.get("id")
    if not vuln_id:
        raise ValueError("NVD item without id")

    descriptions = cve.get("descriptions") or []
    desc_en = next((d.get("value") for d in descriptions if d.get("lang") == "en"), "")
    if not desc_en and descriptions:
        desc_en = descriptions[0].get("value") or ""

    metrics = _extract_cvss(cve.get("metrics") or {})
    score = None
    for blob in (metrics["v31"], metrics["v30"], metrics["v4"], metrics["v2"]):
        if blob and blob.get("score") is not None:
            score = blob["score"]
            break

    weaknesses = []
    for w in cve.get("weaknesses") or []:
        for d in w.get("description") or []:
            if d.get("value"):
                weaknesses.append(d["value"])

    refs = []
    for r in cve.get("references") or []:
        refs.append({"url": r.get("url"), "tags": r.get("tags") or []})

    cpes = []
    for cfg in cve.get("configurations") or []:
        for node in cfg.get("nodes") or []:
            for match in node.get("cpeMatch") or []:
                if match.get("criteria"):
                    cpes.append(match["criteria"])

    defaults = {
        "title": (desc_en or vuln_id)[:512],
        "description_nvd": desc_en or "",
        "severity": severity_from_score(score),
        "cvss_v31": metrics["v31"],
        "cvss_v30": metrics["v30"],
        "cvss_v2": metrics["v2"],
        "cvss_v4": metrics["v4"],
        "cwe": weaknesses,
        "cpe": cpes[:200],
        "references": refs[:100],
        "published_at": _parse_dt(cve.get("published")),
        "modified_at": _parse_dt(cve.get("lastModified")),
        "raw_nvd": item,
    }
    obj, created = Vulnerability.objects.get_or_create(vuln_id=vuln_id, defaults={
        **defaults,
        "source": Vulnerability.Source.NVD,
    })
    if not created:
        for k, v in defaults.items():
            setattr(obj, k, v)
        if obj.source == Vulnerability.Source.BDU:
            obj.source = Vulnerability.Source.MERGED
        elif obj.source != Vulnerability.Source.MERGED:
            obj.source = Vulnerability.Source.NVD
        obj.save()
    return obj


@shared_task(name="apps.vulns.tasks.sync_nvd_incremental")
def sync_nvd_incremental():
    """Incremental NVD API 2.0 sync. Skips if API key missing."""
    s = SystemSettings.get_solo()
    api_key = (s.nvd_api_key or getattr(settings, "NVD_API_KEY", "") or "").strip()
    cp = _checkpoint("nvd")
    if not api_key:
        msg = "NVD API key not configured — skip sync_nvd_incremental"
        logger.info(msg)
        cp.status = SyncCheckpoint.Status.IDLE
        cp.detail = msg
        cp.save(update_fields=["status", "detail", "updated_at"])
        return {"skipped": True, "reason": msg}

    cp.status = SyncCheckpoint.Status.RUNNING
    cp.detail = "starting"
    cp.save(update_fields=["status", "detail", "updated_at"])

    headers = {"apiKey": api_key}
    params = {"resultsPerPage": 100}
    if cp.cursor:
        params["lastModStartDate"] = cp.cursor
        params["lastModEndDate"] = timezone.now().strftime("%Y-%m-%dT%H:%M:%S.000")

    total = 0
    try:
        with httpx.Client(timeout=60.0, headers=headers) as client:
            start_index = 0
            while True:
                params["startIndex"] = start_index
                resp = client.get(NVD_API, params=params)
                resp.raise_for_status()
                data = resp.json()
                vulns = data.get("vulnerabilities") or []
                for item in vulns:
                    _upsert_nvd_item(item)
                    total += 1
                total_results = int(data.get("totalResults") or 0)
                start_index += len(vulns)
                if start_index >= total_results or not vulns:
                    break
                # NVD rate limit courtesy pause for keyed clients is mild; avoid hammering
                if start_index >= 2000:
                    # safety cap per run
                    break

        now = timezone.now()
        cp.cursor = now.strftime("%Y-%m-%dT%H:%M:%S.000")
        cp.last_success_at = now
        cp.status = SyncCheckpoint.Status.SUCCESS
        cp.detail = f"upserted={total}"
        cp.save()
        logger.info("sync_nvd_incremental: upserted %s", total)
        return {"upserted": total}
    except Exception as exc:
        logger.exception("sync_nvd_incremental failed")
        cp.status = SyncCheckpoint.Status.ERROR
        cp.detail = str(exc)[:2000]
        cp.save(update_fields=["status", "detail", "updated_at"])
        return {"error": str(exc)}


@shared_task(name="apps.vulns.tasks.sync_kev")
def sync_kev():
    """Sync CISA KEV catalog into existing CVE rows."""
    s = SystemSettings.get_solo()
    if not s.kev_enabled:
        logger.info("sync_kev: disabled in settings")
        return {"skipped": True}

    url = getattr(settings, "KEV_URL", "") or (
        "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
    )
    cp = _checkpoint("kev")
    cp.status = SyncCheckpoint.Status.RUNNING
    cp.save(update_fields=["status", "updated_at"])
    updated = 0
    try:
        with httpx.Client(timeout=60.0) as client:
            resp = client.get(url)
            resp.raise_for_status()
            data = resp.json()
        for entry in data.get("vulnerabilities") or []:
            cve_id = entry.get("cveID")
            if not cve_id:
                continue
            due = parse_date(entry.get("dueDate") or "") if entry.get("dueDate") else None
            obj, created = Vulnerability.objects.get_or_create(
                vuln_id=cve_id,
                defaults={
                    "title": (entry.get("vulnerabilityName") or cve_id)[:512],
                    "description_nvd": entry.get("shortDescription") or "",
                    "severity": Vulnerability.Severity.HIGH,
                    "is_kev": True,
                    "kev_due_date": due,
                    "source": Vulnerability.Source.NVD,
                    "references": [{"url": entry.get("notes")}] if entry.get("notes") else [],
                },
            )
            if not created:
                obj.is_kev = True
                obj.kev_due_date = due or obj.kev_due_date
                if entry.get("vulnerabilityName") and not obj.title:
                    obj.title = entry["vulnerabilityName"][:512]
                obj.save(update_fields=["is_kev", "kev_due_date", "title", "updated_at"])
            updated += 1
        cp.status = SyncCheckpoint.Status.SUCCESS
        cp.last_success_at = timezone.now()
        cp.detail = f"kev_marked={updated}"
        cp.cursor = data.get("catalogVersion") or ""
        cp.save()
        logger.info("sync_kev: updated %s", updated)
        return {"updated": updated}
    except Exception as exc:
        logger.exception("sync_kev failed")
        cp.status = SyncCheckpoint.Status.ERROR
        cp.detail = str(exc)[:2000]
        cp.save(update_fields=["status", "detail", "updated_at"])
        return {"error": str(exc)}


@shared_task(name="apps.vulns.tasks.sync_bdu")
def sync_bdu():
    """
    Stub BDU sync: if BDU_XLSX_PATH is set, read with openpyxl and merge/upsert.
    Expected columns (flexible): identifier / BDU id, CVE, description, name.
    """
    s = SystemSettings.get_solo()
    if not s.bdu_enabled:
        return {"skipped": True, "reason": "bdu disabled"}

    path = (getattr(settings, "BDU_XLSX_PATH", "") or "").strip()
    cp = _checkpoint("bdu")
    if not path:
        msg = "BDU_XLSX_PATH not set — skip sync_bdu"
        logger.info(msg)
        cp.status = SyncCheckpoint.Status.IDLE
        cp.detail = msg
        cp.save(update_fields=["status", "detail", "updated_at"])
        return {"skipped": True, "reason": msg}

    try:
        from openpyxl import load_workbook
    except ImportError:
        msg = "openpyxl not installed"
        cp.status = SyncCheckpoint.Status.ERROR
        cp.detail = msg
        cp.save(update_fields=["status", "detail", "updated_at"])
        return {"error": msg}

    cp.status = SyncCheckpoint.Status.RUNNING
    cp.save(update_fields=["status", "updated_at"])
    merged = created = 0
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c or "").strip().lower() for c in next(rows, [])]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        i_bdu = col("идентификатор", "bdu", "bdu_id", "id")
        i_cve = col("cve", "cve_id", "identifier")
        i_desc = col("описание", "description", "desc")
        i_name = col("название", "name", "title")

        for row in rows:
            if not row:
                continue
            bdu_id = str(row[i_bdu]).strip() if i_bdu is not None and row[i_bdu] else ""
            cve_id = str(row[i_cve]).strip() if i_cve is not None and row[i_cve] else ""
            desc = str(row[i_desc]).strip() if i_desc is not None and row[i_desc] else ""
            name = str(row[i_name]).strip() if i_name is not None and row[i_name] else ""
            if cve_id.upper().startswith("CVE-"):
                try:
                    cve = Vulnerability.objects.get(vuln_id=cve_id.upper())
                    merge_bdu_into_cve(
                        cve,
                        description_bdu=desc,
                        raw_bdu={"bdu_id": bdu_id, "row": list(row)},
                        title=name,
                    )
                    merged += 1
                except Vulnerability.DoesNotExist:
                    Vulnerability.objects.create(
                        vuln_id=cve_id.upper(),
                        title=(name or cve_id)[:512],
                        description_bdu=desc,
                        description_nvd="",
                        source=Vulnerability.Source.MERGED,
                        raw_bdu={"bdu_id": bdu_id},
                        severity=Vulnerability.Severity.UNKNOWN,
                    )
                    created += 1
            elif bdu_id:
                vid = bdu_id if bdu_id.upper().startswith("BDU") else f"BDU:{bdu_id}"
                _, was_created = Vulnerability.objects.get_or_create(
                    vuln_id=vid,
                    defaults={
                        "title": (name or vid)[:512],
                        "description_bdu": desc,
                        "source": Vulnerability.Source.BDU,
                        "raw_bdu": {"bdu_id": bdu_id},
                        "severity": Vulnerability.Severity.UNKNOWN,
                    },
                )
                if was_created:
                    created += 1
                else:
                    merged += 1

        cp.status = SyncCheckpoint.Status.SUCCESS
        cp.last_success_at = timezone.now()
        cp.detail = f"merged={merged} created={created}"
        cp.cursor = path
        cp.save()
        return {"merged": merged, "created": created}
    except Exception as exc:
        logger.exception("sync_bdu failed")
        cp.status = SyncCheckpoint.Status.ERROR
        cp.detail = str(exc)[:2000]
        cp.save(update_fields=["status", "detail", "updated_at"])
        return {"error": str(exc)}
